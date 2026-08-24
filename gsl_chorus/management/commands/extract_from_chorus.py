import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from gsl_chorus.models import LIGNE_IDENTITY_FIELDS, SuiviFinancier
from gsl_demarches_simplifiees.models import Dossier
from gsl_projet.constants import DOTATION_DETR, DOTATION_DSIL

COLUMN_HEADERS = {
    "ej": "Nº pce réf",
    "dn": "Axe ministériel 2",
    "domaine_fonctionnel": "Dom. fonc.",
    "montant": "Mont.DT",
    "date": "CB dteComp",
    "date_engagement": "Saisie le",
    "tv": "TV",
    "type_montant": "TypeMtant",
    "poste": "Pos.r",
    "compte_general": "Cpte gén.",
}

DOMAINE_FONCTIONNEL_DOTATION = {
    "0119-01-06": DOTATION_DETR,
    "0119-01-07": DOTATION_DSIL,
}

# The DN is hand-typed in the "Axe ministériel 2" column in inconsistent formats
# ("DN-30736671", "DS : 15480250", "DS-n°16090518", bare "15345041"…) mixed with
# noise ("CRTE-2022", "DOSSIER PAPIER"…). Valid DS/DN numbers are exactly 8 digits
# (like Dossier.ds_number), so we match an 8-digit run with no digit on either side.
DN_RE = re.compile(r"(?<!\d)\d{8}(?!\d)")

BATCH_SIZE = 2000


def extract_dn_number(value):
    if not value:
        return None
    match = DN_RE.search(str(value))
    return int(match.group()) if match else None


def parse_montant(value):
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.strptime(text, "%d.%m.%Y").date()
    except ValueError:
        return None


class Command(BaseCommand):
    """
    uv run python manage.py extract_from_chorus <fichier.xlsx>
    """

    help = (
        "Importe un export « Journal des pièces » de Chorus (XLSX) dans SuiviFinancier."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "file", help="Chemin vers le fichier XLSX exporté de Chorus"
        )

    def handle(self, *args, **options):
        self.import_file(options["file"])

    def iter_records(self, path):
        """Yield one dict of raw cell values per data row, keyed like COLUMN_HEADERS."""
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Fichier introuvable : {path}")
        sheet = workbook.active
        columns, header_row = self.locate_columns(sheet)
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            record = {key: row[index] for key, index in columns.items()}
            ej = record["ej"]
            dn = record["dn"]
            if (ej is None or not str(ej).strip()) and (
                dn is None or not str(dn).strip()
            ):
                continue
            yield record
        workbook.close()

    def locate_columns(self, sheet):
        for header_row, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1
        ):
            headers = {
                (str(v).strip() if v is not None else ""): i for i, v in enumerate(row)
            }
            columns = {
                key: headers[label]
                for key, label in COLUMN_HEADERS.items()
                if label in headers
            }
            if len(columns) == len(COLUMN_HEADERS):
                return columns, header_row
        missing = ", ".join(COLUMN_HEADERS.values())
        raise CommandError(
            f"En-têtes introuvables dans les 20 premières lignes (attendus : {missing})."
        )

    @transaction.atomic
    def import_file(self, path):
        known = set(Dossier.objects.values_list("ds_number", flat=True))
        seen = set(SuiviFinancier.objects.values_list(*LIGNE_IDENTITY_FIELDS))
        batch = []
        imported = skipped = duplicate = matched = 0
        for record in self.iter_records(path):
            ej = str(record["ej"]).strip() if record["ej"] is not None else ""
            montant = parse_montant(record["montant"])
            if not ej or montant is None:
                skipped += 1
                continue
            fields = {
                "ej": ej,
                "dn": extract_dn_number(record["dn"]),
                "dotation": DOMAINE_FONCTIONNEL_DOTATION.get(
                    str(record["domaine_fonctionnel"] or "").strip(), ""
                ),
                "montant": montant,
                "date_transaction": parse_date(record["date"]),
                "date_engagement": parse_date(record["date_engagement"]),
                "tv": str(record["tv"] or "").strip()[:5],
                "type_montant": str(record["type_montant"] or "").strip()[:10],
                "poste": str(record["poste"] or "").strip()[:10],
                "compte_general": str(record["compte_general"] or "").strip()[:20],
            }
            key = tuple(fields[name] for name in LIGNE_IDENTITY_FIELDS)
            if key in seen:
                duplicate += 1
                continue
            seen.add(key)
            batch.append(SuiviFinancier(**fields))
            imported += 1
            if fields["dn"] in known:
                matched += 1
            if len(batch) >= BATCH_SIZE:
                SuiviFinancier.objects.bulk_create(batch, ignore_conflicts=True)
                batch = []
        if batch:
            SuiviFinancier.objects.bulk_create(batch, ignore_conflicts=True)

        self.stdout.write(self.style.SUCCESS(f"{imported} nouvelle(s) ligne(s)."))
        self.stdout.write(f"{matched} rattachée(s) à un dossier connu (DN).")
        if duplicate:
            self.stdout.write(f"{duplicate} déjà présente(s), ignorée(s).")
        if skipped:
            self.stdout.write(
                f"{skipped} ligne(s) ignorée(s) (sans EJ ou sans montant)."
            )
