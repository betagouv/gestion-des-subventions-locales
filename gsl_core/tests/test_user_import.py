from io import BytesIO

import openpyxl
import pytest
import tablib
from django.urls import reverse

from gsl_core.admin import CollegueAdmin
from gsl_core.models import Collegue, Perimetre
from gsl_core.resources import CollegueResource

from .factories import (
    ArrondissementFactory,
    ClientWithLoggedUserFactory,
    CollegueFactory,
    DepartementFactory,
    PerimetreArrondissementFactory,
    PerimetreDepartementalFactory,
    RegionFactory,
)

pytestmark = pytest.mark.django_db


@pytest.fixture
def geographic_data():
    """Setup geographic data for tests"""
    region = RegionFactory(insee_code="84", name="Auvergne-Rhône-Alpes")
    dept_ain = DepartementFactory(insee_code="01", name="AIN", region=region)
    arr_belley = ArrondissementFactory(
        insee_code="011", name="Belley", departement=dept_ain
    )

    # Create perimeters
    perimetre_dept = PerimetreDepartementalFactory(
        departement=dept_ain, region=region, arrondissement=None
    )
    perimetre_arr = PerimetreArrondissementFactory(
        departement=dept_ain, region=region, arrondissement=arr_belley
    )

    # Also create AUBE department (used in sample Excel file)
    region_ge = RegionFactory(insee_code="44", name="Grand Est")
    dept_aube = DepartementFactory(insee_code="10", name="AUBE", region=region_ge)

    return {
        "region": region,
        "dept": dept_ain,
        "arr": arr_belley,
        "perimetre_dept": perimetre_dept,
        "perimetre_arr": perimetre_arr,
        "dept_aube": dept_aube,
    }


@pytest.fixture
def sample_excel_file():
    """Create a basic Excel file with hierarchical user data"""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Row 1: empty
    ws.append([None, None, None, None, None, None, None])

    # Row 2: Headers
    ws.append(
        [
            None,
            "Département",
            "Périmètre",
            "Code arrondissement",
            "Contacts",
            "Prénom",
            "Nom",
        ]
    )

    # Row 3-5: Department section for AIN (01)
    ws.append(["01", "AIN", None, None, "pref@ain.gouv.fr", "Jean-Claude", "Dupont"])
    ws.append([None, None, None, None, "user1@ain.gouv.fr", "Marie", "De La Fontaine"])
    ws.append(
        [None, None, None, None, "user2@ain.gouv.fr", "Pierre Jean", "Durand-Martin"]
    )

    # Row 6-7: Arrondissement section (Belley - 011)
    ws.append(
        [
            None,
            None,
            "SP de Belley",
            "011",
            "arr-user1@ain.gouv.fr",
            "Sophie",
            "Bernard",
        ]
    )
    ws.append([None, None, None, None, "arr-user2@ain.gouv.fr", "Luc", "Petit"])

    # Row for another department
    ws.append(["10", "AUBE", None, None, "pref@aube.gouv.fr", "Alice", "Robert"])

    # Save to BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@pytest.fixture
def admin_client(geographic_data):
    """Create an authenticated admin client"""
    admin_user = CollegueFactory(is_staff=True, is_superuser=True)
    return ClientWithLoggedUserFactory(user=admin_user)


class TestExcelUserDataParser:
    def test_parse_hierarchical_structure(self, sample_excel_file, geographic_data):
        """Test parser correctly handles department/arrondissement state tracking and name parsing"""
        admin = CollegueAdmin(Collegue, None)
        dataset = admin.parse_excel_to_dataset(sample_excel_file)

        assert len(dataset) == 6

        # Department-level users (AIN - 01)
        assert dataset[0] == ("pref@ain.gouv.fr", "01", None, "Jean-Claude", "Dupont")
        assert dataset[1] == (
            "user1@ain.gouv.fr",
            "01",
            None,
            "Marie",
            "De La Fontaine",
        )
        assert dataset[2] == (
            "user2@ain.gouv.fr",
            "01",
            None,
            "Pierre Jean",
            "Durand-Martin",
        )

        # Arrondissement-level users (Belley - 011)
        assert dataset[3] == ("arr-user1@ain.gouv.fr", "01", "011", "Sophie", "Bernard")
        assert dataset[4] == ("arr-user2@ain.gouv.fr", "01", "011", "Luc", "Petit")

        # Department-level user (AUBE - 10)
        assert dataset[5] == ("pref@aube.gouv.fr", "10", None, "Alice", "Robert")

    def test_parse_normalizes_department_codes(self):
        """Test department code normalization (1 -> '01', 10 -> '10')"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, None, None, None, None])  # Row 1: empty
        ws.append(
            [None, "Département", "Périmètre", "Code arrondissement", "Contacts"]
        )  # Row 2: headers
        ws.append([1, "TEST", None, None, "test@example.com"])  # 1 -> '01'
        ws.append([10, "TEST2", None, None, "test2@example.com"])  # 10 -> '10'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        admin = CollegueAdmin(Collegue, None)
        dataset = admin.parse_excel_to_dataset(buffer)

        assert len(dataset) == 2
        assert dataset[0][1] == "01"  # Normalized to '01'
        assert dataset[1][1] == "10"  # Stays '10'

    def test_parse_handles_corsica_codes(self):
        """Test special Corsica department codes 2A and 2B"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, None, None, None, None])  # Row 1: empty
        ws.append(
            [None, "Département", "Périmètre", "Code arrondissement", "Contacts"]
        )  # Row 2: headers
        ws.append(["2A", "Corse-du-Sud", None, None, "test@corse.fr"])
        ws.append(["2B", "Haute-Corse", None, None, "test2@corse.fr"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        admin = CollegueAdmin(Collegue, None)
        dataset = admin.parse_excel_to_dataset(buffer)

        assert len(dataset) == 2
        assert dataset[0][1] == "2A"
        assert dataset[1][1] == "2B"

    def test_parse_handles_float_department_codes(self):
        """Test that float department codes from Excel are converted to integers"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, None, None, None, None])  # Row 1: empty
        ws.append(
            [None, "Département", "Périmètre", "Code arrondissement", "Contacts"]
        )  # Row 2: headers

        # Simulate openpyxl returning floats for numeric cells
        ws.append([10.0, "AUBE", None, None, "test@aube.fr"])  # Float value
        ws.append([1.0, "AIN", None, None, "test@ain.fr"])  # Float requiring padding

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        admin = CollegueAdmin(Collegue, None)
        dataset = admin.parse_excel_to_dataset(buffer)

        assert len(dataset) == 2
        assert dataset[0][1] == "10"  # Not "10.0"
        assert dataset[1][1] == "01"  # Padded single digit

    def test_parse_handles_float_arrondissement_codes(self):
        """Test that float arrondissement codes from Excel are converted and padded"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, None, None, None, None, None, None])  # Row 1: empty
        ws.append(
            [
                None,
                "Département",
                "Périmètre",
                "Code arrondissement",
                "Contacts",
                "Prénom",
                "Nom",
            ]
        )  # Row 2: headers

        # Department section
        ws.append(["01", "AIN", None, None, "dept@ain.fr", "Jean", "Test"])
        # Arrondissement with float code (Excel stores '011' as 11.0)
        ws.append([None, None, "SP de Belley", 11.0, "arr@ain.fr", "Marie", "Test"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        admin = CollegueAdmin(Collegue, None)
        dataset = admin.parse_excel_to_dataset(buffer)

        assert len(dataset) == 2
        assert dataset[0] == ("dept@ain.fr", "01", None, "Jean", "Test")  # Dept level
        assert dataset[1] == (
            "arr@ain.fr",
            "01",
            "011",
            "Marie",
            "Test",
        )  # Arr code normalized to 011

    def test_parse_skips_invalid_emails(self):
        """Test that rows without valid emails are skipped"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, None, None, None, None])  # Row 1: empty
        ws.append(
            [None, "Département", "Périmètre", "Code arrondissement", "Contacts"]
        )  # Row 2: headers
        ws.append(["01", "AIN", None, None, "valid@example.com"])
        ws.append([None, None, None, None, "not-an-email"])  # No @
        ws.append([None, None, None, None, None])  # Empty
        ws.append([None, None, None, None, "another@valid.com"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        admin = CollegueAdmin(Collegue, None)
        dataset = admin.parse_excel_to_dataset(buffer)

        assert len(dataset) == 2
        assert dataset[0][0] == "valid@example.com"
        assert dataset[1][0] == "another@valid.com"

    def test_parse_skips_emails_without_department_context(self):
        """Test that emails before any department section are skipped"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, None, None, None, None])  # Row 1: empty
        ws.append(
            [None, "Département", "Périmètre", "Code arrondissement", "Contacts"]
        )  # Row 2: headers
        ws.append([None, None, None, None, "orphan@example.com"])  # No dept yet
        ws.append(["01", "AIN", None, None, "valid@ain.gouv.fr"])
        ws.append([None, None, None, None, "another@ain.gouv.fr"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        admin = CollegueAdmin(Collegue, None)
        dataset = admin.parse_excel_to_dataset(buffer)

        assert len(dataset) == 2
        assert "orphan@example.com" not in [row[0] for row in dataset]
        assert dataset[0][0] == "valid@ain.gouv.fr"

    def test_parse_state_resets_on_new_department(self):
        """Test that arrondissement state resets when new department starts"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, None, None, None, None, None, None])  # Row 1: empty
        ws.append(
            [
                None,
                "Département",
                "Périmètre",
                "Code arrondissement",
                "Contacts",
                "Prénom",
                "Nom",
            ]
        )  # Row 2: headers

        # First department with arrondissement
        ws.append(["01", "AIN", None, None, "dept1@ain.fr", "Jean", "Dupont"])
        ws.append([None, None, "SP de Belley", "011", "arr@ain.fr", "Marie", "Martin"])
        ws.append(
            [None, None, None, None, "arr2@ain.fr", "Pierre", "Durand"]
        )  # Still in arr context

        # Second department - should reset arr context
        ws.append(["10", "AUBE", None, None, "dept2@aube.fr", "Sophie", "Bernard"])
        ws.append(
            [None, None, None, None, "dept2-user@aube.fr", "Luc", "Petit"]
        )  # Back to dept context

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        admin = CollegueAdmin(Collegue, None)
        dataset = admin.parse_excel_to_dataset(buffer)

        assert len(dataset) == 5

        # First dept-level
        assert dataset[0] == ("dept1@ain.fr", "01", None, "Jean", "Dupont")
        # Arr-level
        assert dataset[1] == ("arr@ain.fr", "01", "011", "Marie", "Martin")
        assert dataset[2] == ("arr2@ain.fr", "01", "011", "Pierre", "Durand")
        # Second dept-level (arr context reset)
        assert dataset[3] == ("dept2@aube.fr", "10", None, "Sophie", "Bernard")
        assert dataset[4] == ("dept2-user@aube.fr", "10", None, "Luc", "Petit")

    def test_parse_handles_names_with_dashes_and_spaces(self):
        """Test that names with dashes and spaces are correctly parsed"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, None, None, None, None, None, None])  # Row 1: empty
        ws.append(
            [
                None,
                "Département",
                "Périmètre",
                "Code arrondissement",
                "Contacts",
                "Prénom",
                "Nom",
            ]
        )  # Row 2: headers

        # Test various name formats with dashes and spaces
        ws.append(
            ["01", "AIN", None, None, "user1@ain.fr", "Jean-Claude", "Dupont"]
        )  # First name with dash
        ws.append(
            [None, None, None, None, "user2@ain.fr", "Marie-France", "De La Fontaine"]
        )  # Both with dash/space
        ws.append(
            [None, None, None, None, "user3@ain.fr", "Pierre Jean", "Durand-Martin"]
        )  # First name with space, last name with dash
        ws.append(
            [None, None, None, None, "user4@ain.fr", "Anne Marie", "Van Der Berg"]
        )  # Both with spaces
        ws.append(
            [
                None,
                None,
                None,
                None,
                "user5@ain.fr",
                "Jean-Pierre Louis",
                "De Saint-Exupéry",
            ]
        )  # Complex names with both

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        admin = CollegueAdmin(Collegue, None)
        dataset = admin.parse_excel_to_dataset(buffer)

        assert len(dataset) == 5

        # Verify all names are preserved exactly
        assert dataset[0] == ("user1@ain.fr", "01", None, "Jean-Claude", "Dupont")
        assert dataset[1] == (
            "user2@ain.fr",
            "01",
            None,
            "Marie-France",
            "De La Fontaine",
        )
        assert dataset[2] == (
            "user3@ain.fr",
            "01",
            None,
            "Pierre Jean",
            "Durand-Martin",
        )
        assert dataset[3] == (
            "user4@ain.fr",
            "01",
            None,
            "Anne Marie",
            "Van Der Berg",
        )
        assert dataset[4] == (
            "user5@ain.fr",
            "01",
            None,
            "Jean-Pierre Louis",
            "De Saint-Exupéry",
        )


@pytest.mark.django_db
class TestCollegueResource:
    def test_import_creates_new_users_department_level(self, geographic_data):
        """Test resource creates users with department-level perimetre"""
        dataset = tablib.Dataset()
        dataset.headers = [
            "email",
            "departement_code",
            "arrondissement_code",
            "first_name",
            "last_name",
        ]
        dataset.append(["user1@ain.gouv.fr", "01", None, "Jean", "Dupont"])
        dataset.append(["user2@ain.gouv.fr", "01", None, "Marie", "Martin"])

        resource = CollegueResource()
        result = resource.import_data(dataset, dry_run=False)

        assert result.totals["new"] == 2
        assert result.totals["error"] == 0

        # Verify users created
        user1 = Collegue.objects.get(email="user1@ain.gouv.fr")
        assert user1.username == "user1@ain.gouv.fr"
        assert user1.is_active is True
        assert user1.is_staff is False
        assert not user1.has_usable_password()
        assert user1.first_name == "Jean"
        assert user1.last_name == "Dupont"

        user2 = Collegue.objects.get(email="user2@ain.gouv.fr")
        assert user2.first_name == "Marie"
        assert user2.last_name == "Martin"

        # Verify perimetre
        assert user1.perimetre.departement.insee_code == "01"
        assert user1.perimetre.arrondissement is None

    def test_import_creates_new_users_arrondissement_level(self, geographic_data):
        """Test resource creates users with arrondissement-level perimetre"""
        dataset = tablib.Dataset()
        dataset.headers = [
            "email",
            "departement_code",
            "arrondissement_code",
            "first_name",
            "last_name",
        ]
        dataset.append(["user@ain.gouv.fr", "01", "011", "Sophie", "Bernard"])

        resource = CollegueResource()
        result = resource.import_data(dataset, dry_run=False)

        assert result.totals["new"] == 1

        user = Collegue.objects.get(email="user@ain.gouv.fr")
        assert user.perimetre.arrondissement.insee_code == "011"
        assert user.perimetre.departement.insee_code == "01"

    def test_import_updates_existing_users(self, geographic_data):
        """Test resource updates existing user's perimetre"""
        # Create user with arrondissement perimetre (will be updated to dept perimetre)
        old_perimetre = geographic_data["perimetre_arr"]
        existing_user = CollegueFactory(
            email="existing@ain.gouv.fr",
            username="old_username",
            perimetre=old_perimetre,
        )

        # Import should update perimetre from arrondissement to department level
        dataset = tablib.Dataset()
        dataset.headers = [
            "email",
            "departement_code",
            "arrondissement_code",
            "first_name",
            "last_name",
        ]
        dataset.append(["existing@ain.gouv.fr", "01", None, "Jean", "Dupont"])

        resource = CollegueResource()
        result = resource.import_data(dataset, dry_run=False)

        assert result.totals["update"] == 1
        assert result.totals["new"] == 0

        existing_user.refresh_from_db()
        assert existing_user.perimetre.departement.insee_code == "01"
        assert existing_user.perimetre == geographic_data["perimetre_dept"]
        assert existing_user.perimetre != old_perimetre

    def test_import_sets_unusable_password_for_new_users_only(self, geographic_data):
        """Test unusable password only set for new users, not updates"""
        # Create existing user with usable password
        existing_user = CollegueFactory(
            email="existing@ain.gouv.fr", perimetre=geographic_data["perimetre_dept"]
        )
        existing_user.set_password("testpassword123")
        existing_user.save()

        # Import both new and existing
        dataset = tablib.Dataset()
        dataset.headers = [
            "email",
            "departement_code",
            "arrondissement_code",
            "first_name",
            "last_name",
        ]
        dataset.append(["existing@ain.gouv.fr", "01", None, "Jean", "Dupont"])
        dataset.append(["newuser@ain.gouv.fr", "01", None, "Marie", "Martin"])

        resource = CollegueResource()
        result = resource.import_data(dataset, dry_run=False)

        assert result.totals["new"] == 1
        assert result.totals["update"] == 1

        # Existing user should keep usable password
        existing_user.refresh_from_db()
        assert existing_user.has_usable_password()

        # New user should have unusable password
        new_user = Collegue.objects.get(email="newuser@ain.gouv.fr")
        assert not new_user.has_usable_password()

    def test_import_creates_perimetres_automatically(self, geographic_data):
        """Test Perimetre objects are created/found correctly"""
        initial_count = Perimetre.objects.count()

        dataset = tablib.Dataset()
        dataset.headers = [
            "email",
            "departement_code",
            "arrondissement_code",
            "first_name",
            "last_name",
        ]
        dataset.append(["user1@ain.gouv.fr", "01", None, "Jean", "Dupont"])
        dataset.append(
            ["user2@ain.gouv.fr", "01", None, "Marie", "Martin"]
        )  # Same perimetre
        dataset.append(
            ["user3@ain.gouv.fr", "01", "011", "Pierre", "Durand"]
        )  # Different perimetre

        resource = CollegueResource()
        resource.import_data(dataset, dry_run=False)

        # Perimetres already exist from fixture, so they're reused (no new ones created)
        assert Perimetre.objects.count() == initial_count

        # Verify users share perimetre when appropriate
        user1 = Collegue.objects.get(email="user1@ain.gouv.fr")
        user2 = Collegue.objects.get(email="user2@ain.gouv.fr")
        assert user1.perimetre == user2.perimetre

    def test_import_handles_missing_department(self):
        """Test error when department not found in database"""
        dataset = tablib.Dataset()
        dataset.headers = [
            "email",
            "departement_code",
            "arrondissement_code",
            "first_name",
            "last_name",
        ]
        dataset.append(
            ["test@example.com", "99", None, "Jean", "Dupont"]
        )  # Non-existent dept

        resource = CollegueResource()
        result = resource.import_data(dataset, dry_run=True)

        assert result.has_errors()
        error_msg = str(result.rows[0].errors[0].error)
        assert "Departement" in error_msg and (
            "not found" in error_msg or "does not exist" in error_msg
        )

    def test_import_handles_missing_arrondissement(self):
        """Test error when arrondissement not found in database"""
        dataset = tablib.Dataset()
        dataset.headers = [
            "email",
            "departement_code",
            "arrondissement_code",
            "first_name",
            "last_name",
        ]
        dataset.append(
            ["test@example.com", "01", "999", "Jean", "Dupont"]
        )  # Non-existent arr

        resource = CollegueResource()
        result = resource.import_data(dataset, dry_run=True)

        assert result.has_errors()
        error_msg = str(result.rows[0].errors[0].error)
        assert "Arrondissement" in error_msg and (
            "not found" in error_msg or "does not exist" in error_msg
        )

    def test_import_handles_missing_department_code(self):
        """Test error when department_code is missing"""
        dataset = tablib.Dataset()
        dataset.headers = [
            "email",
            "departement_code",
            "arrondissement_code",
            "first_name",
            "last_name",
        ]
        dataset.append(["test@example.com", None, None, "Jean", "Dupont"])

        resource = CollegueResource()
        result = resource.import_data(dataset, dry_run=True)

        assert result.has_errors()
        assert "Missing departement_code" in str(result.rows[0].errors[0].error)

    def test_import_dry_run_does_not_create_users(self, geographic_data):
        """Test dry_run mode doesn't actually create users"""
        initial_count = Collegue.objects.count()

        dataset = tablib.Dataset()
        dataset.headers = [
            "email",
            "departement_code",
            "arrondissement_code",
            "first_name",
            "last_name",
        ]
        dataset.append(["user@ain.gouv.fr", "01", None, "Jean", "Dupont"])

        resource = CollegueResource()
        result = resource.import_data(dataset, dry_run=True)

        assert result.totals["new"] == 1  # Shows what would be created
        assert Collegue.objects.count() == initial_count  # But doesn't actually create

    def test_import_handles_names_with_dashes_and_spaces(self, geographic_data):
        """Test resource correctly imports names with dashes and spaces"""
        dataset = tablib.Dataset()
        dataset.headers = [
            "email",
            "departement_code",
            "arrondissement_code",
            "first_name",
            "last_name",
        ]
        dataset.append(["user1@ain.gouv.fr", "01", None, "Jean-Claude", "Dupont"])
        dataset.append(
            ["user2@ain.gouv.fr", "01", None, "Marie-France", "De La Fontaine"]
        )
        dataset.append(
            ["user3@ain.gouv.fr", "01", None, "Pierre Jean", "Durand-Martin"]
        )
        dataset.append(["user4@ain.gouv.fr", "01", None, "Anne Marie", "Van Der Berg"])

        resource = CollegueResource()
        result = resource.import_data(dataset, dry_run=False)

        assert result.totals["new"] == 4
        assert result.totals["error"] == 0

        # Verify names are preserved exactly
        user1 = Collegue.objects.get(email="user1@ain.gouv.fr")
        assert user1.first_name == "Jean-Claude"
        assert user1.last_name == "Dupont"

        user2 = Collegue.objects.get(email="user2@ain.gouv.fr")
        assert user2.first_name == "Marie-France"
        assert user2.last_name == "De La Fontaine"

        user3 = Collegue.objects.get(email="user3@ain.gouv.fr")
        assert user3.first_name == "Pierre Jean"
        assert user3.last_name == "Durand-Martin"

        user4 = Collegue.objects.get(email="user4@ain.gouv.fr")
        assert user4.first_name == "Anne Marie"
        assert user4.last_name == "Van Der Berg"


@pytest.mark.django_db
class TestIntegration:
    def test_full_excel_to_users_flow(self, sample_excel_file, geographic_data):
        """Integration test: Excel file -> Parser -> Resource -> Users created"""
        # Step 1: Parse Excel
        admin = CollegueAdmin(Collegue, None)
        dataset = admin.parse_excel_to_dataset(sample_excel_file)
        assert len(dataset) > 0

        # Step 2: Import via resource
        resource = CollegueResource()
        result = resource.import_data(dataset, dry_run=False)

        assert result.totals["new"] > 0
        assert result.totals["error"] == 0

        # Step 3: Verify users exist with correct perimetres and names
        user = Collegue.objects.get(email="pref@ain.gouv.fr")
        assert user.perimetre.departement.insee_code == "01"
        assert user.perimetre.arrondissement is None
        assert user.first_name == "Jean-Claude"
        assert user.last_name == "Dupont"

        arr_user = Collegue.objects.get(email="arr-user1@ain.gouv.fr")
        assert arr_user.perimetre.arrondissement.insee_code == "011"
        assert arr_user.first_name == "Sophie"
        assert arr_user.last_name == "Bernard"

    def test_re_import_updates_instead_of_duplicating(
        self, sample_excel_file, geographic_data
    ):
        """Test re-importing same file doesn't create duplicates"""
        admin = CollegueAdmin(Collegue, None)
        resource = CollegueResource()

        # First import
        dataset1 = admin.parse_excel_to_dataset(sample_excel_file)
        result1 = resource.import_data(dataset1, dry_run=False)
        created_count = result1.totals["new"]
        assert created_count > 0  # Ensure users were created

        initial_user_count = Collegue.objects.count()

        # Reset file pointer
        sample_excel_file.seek(0)

        # Second import - should not create duplicates
        # Note: With skip_unchanged=True, unchanged rows are skipped (not counted as updates)
        dataset2 = admin.parse_excel_to_dataset(sample_excel_file)
        result2 = resource.import_data(dataset2, dry_run=False)

        # Main assertion: no new users created
        assert result2.totals["new"] == 0
        assert Collegue.objects.count() == initial_user_count  # No duplicates


class TestCollegueAdminImportFormDisplay:
    """Test what's displayed on the import form page"""

    def test_import_form_hides_format_field(self, admin_client, geographic_data):
        """Test that format field is hidden from the form"""
        url = reverse("admin:gsl_core_collegue_import")
        response = admin_client.get(url)

        assert response.status_code == 200
        content = response.content.decode("utf-8")

        # Format field should be present but hidden
        assert 'name="format"' in content
        assert 'type="hidden"' in content

        # Format label "Format :" should not be visible (French colon with space)
        assert "Format :" not in content
        # Also check without space (English style)
        assert ">Format:<" not in content

    def test_import_form_hides_resource_fields_message(
        self, admin_client, geographic_data
    ):
        """Test that 'This importer will import the following fields' is hidden"""
        url = reverse("admin:gsl_core_collegue_import")
        response = admin_client.get(url)

        assert response.status_code == 200
        content = response.content.decode("utf-8")

        # French message should not appear
        assert "Cet importateur va importer les champs suivants" not in content

        # English message should not appear
        assert "This importer will import the following fields" not in content

    def test_import_form_shows_only_file_upload(self, admin_client, geographic_data):
        """Test that the form only shows file upload field"""
        url = reverse("admin:gsl_core_collegue_import")
        response = admin_client.get(url)

        assert response.status_code == 200
        content = response.content.decode("utf-8")

        # File upload field should be present
        assert 'name="import_file"' in content
        assert 'type="file"' in content

        # Submit button should be present
        assert 'type="submit"' in content


class TestCollegueAdminExcelImportEndToEnd:
    """End-to-end test for Excel import through Django admin"""

    def test_excel_import_creates_users_in_database(
        self, admin_client, sample_excel_file, geographic_data
    ):
        """Test complete Excel import flow: upload file -> confirm -> verify DB"""
        import re

        # Count existing users
        initial_count = Collegue.objects.count()

        # Step 1: Upload Excel file to import form
        url = reverse("admin:gsl_core_collegue_import")

        with open("/tmp/test_users.xlsx", "wb") as f:
            f.write(sample_excel_file.read())

        with open("/tmp/test_users.xlsx", "rb") as excel_file:
            response = admin_client.post(
                url,
                {
                    "format": "0",  # CSV format (index 0 from available formats)
                    "import_file": excel_file,
                },
            )

        # Should get confirmation page (200)
        assert response.status_code == 200

        # Step 2: Extract all form data from confirmation page
        content = response.content.decode("utf-8")

        # Extract all hidden form fields
        import_file_name_match = re.search(
            r'name="import_file_name" value="([^"]+)"', content
        )
        assert import_file_name_match, "Could not find import_file_name in response"

        original_file_name_match = re.search(
            r'name="original_file_name" value="([^"]+)"', content
        )
        format_match = re.search(r'name="format" value="([^"]+)"', content)

        # Step 3: Submit confirmation with all required fields
        confirm_url = reverse("admin:gsl_core_collegue_process_import")
        confirm_data = {
            "import_file_name": import_file_name_match.group(1),
            "format": format_match.group(1) if format_match else "0",
        }

        # Add original_file_name if present
        if original_file_name_match:
            confirm_data["original_file_name"] = original_file_name_match.group(1)

        response = admin_client.post(confirm_url, confirm_data)

        # Import should succeed (returns 302 redirect)
        assert response.status_code == 302

        # Step 4: Verify users were created in database
        final_count = Collegue.objects.count()

        # Should have 5 new users (6 in file - 1 AUBE that may not have perimetre)
        assert final_count >= initial_count + 5

        # Verify specific users exist with correct data
        # Department-level users
        pref_user = Collegue.objects.get(email="pref@ain.gouv.fr")
        assert pref_user.perimetre == geographic_data["perimetre_dept"]
        assert pref_user.first_name == "Jean-Claude"
        assert pref_user.last_name == "Dupont"

        user1 = Collegue.objects.get(email="user1@ain.gouv.fr")
        assert user1.perimetre == geographic_data["perimetre_dept"]
        assert user1.first_name == "Marie"
        assert user1.last_name == "De La Fontaine"

        user2 = Collegue.objects.get(email="user2@ain.gouv.fr")
        assert user2.perimetre == geographic_data["perimetre_dept"]

        # Arrondissement-level users
        arr_user1 = Collegue.objects.get(email="arr-user1@ain.gouv.fr")
        assert arr_user1.perimetre == geographic_data["perimetre_arr"]
        assert arr_user1.first_name == "Sophie"
        assert arr_user1.last_name == "Bernard"

        arr_user2 = Collegue.objects.get(email="arr-user2@ain.gouv.fr")
        assert arr_user2.perimetre == geographic_data["perimetre_arr"]

    def test_excel_import_with_simple_file(self, admin_client, geographic_data):
        """Test Excel import with a minimal valid file"""
        import re

        # Count existing users
        initial_count = Collegue.objects.count()

        # Create a simple Excel file with just 2 users
        wb = openpyxl.Workbook()
        ws = wb.active

        # Row 1: empty
        ws.append([None, None, None, None, None, None, None])

        # Row 2: Headers
        ws.append(
            [
                None,
                "Département",
                "Périmètre",
                "Code arrondissement",
                "Contacts",
                "Prénom",
                "Nom",
            ]
        )

        # Row 3-4: Department users for AIN (01)
        ws.append(["01", "AIN", None, None, "simple1@ain.gouv.fr", "Paul", "Lefebvre"])
        ws.append([None, None, None, None, "simple2@ain.gouv.fr", "Claire", "Moreau"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Upload and import
        url = reverse("admin:gsl_core_collegue_import")

        with open("/tmp/simple_users.xlsx", "wb") as f:
            f.write(buffer.read())

        with open("/tmp/simple_users.xlsx", "rb") as excel_file:
            response = admin_client.post(
                url,
                {
                    "format": "0",  # CSV format (index 0 from available formats)
                    "import_file": excel_file,
                },
            )

        # Should get confirmation page
        assert response.status_code == 200

        # Extract form data from confirmation page
        content = response.content.decode("utf-8")

        # Extract all hidden form fields
        import_file_name_match = re.search(
            r'name="import_file_name" value="([^"]+)"', content
        )
        assert import_file_name_match, "Could not find import_file_name in response"

        original_file_name_match = re.search(
            r'name="original_file_name" value="([^"]+)"', content
        )
        format_match = re.search(r'name="format" value="([^"]+)"', content)

        # Submit confirmation with all required fields
        confirm_url = reverse("admin:gsl_core_collegue_process_import")
        confirm_data = {
            "import_file_name": import_file_name_match.group(1),
            "format": format_match.group(1) if format_match else "0",
        }

        # Add original_file_name if present
        if original_file_name_match:
            confirm_data["original_file_name"] = original_file_name_match.group(1)

        response = admin_client.post(confirm_url, confirm_data)

        # Import should succeed (returns 302 redirect)
        assert response.status_code == 302

        # Verify 2 new users were created
        final_count = Collegue.objects.count()
        assert final_count == initial_count + 2

        # Verify users exist with correct perimetre and names
        user1 = Collegue.objects.get(email="simple1@ain.gouv.fr")
        assert user1.perimetre == geographic_data["perimetre_dept"]
        assert user1.first_name == "Paul"
        assert user1.last_name == "Lefebvre"

        user2 = Collegue.objects.get(email="simple2@ain.gouv.fr")
        assert user2.perimetre == geographic_data["perimetre_dept"]
        assert user2.first_name == "Claire"
        assert user2.last_name == "Moreau"
